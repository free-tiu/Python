from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook

from app01.models import Department
from app01.utils.pagination import Pagination


def depart_list(request):
    """ 部门列表 """
    depart_list = Department.objects.all().count()

    # 获取部门表中所有部门信息 并 添加分页
    queryset = Department.objects.all()
    # 实例化分页组件
    page_object = Pagination(request, queryset, page_size=10)
    page_queryset = page_object.page_queryset
    pageNums = page_object.html()

    context = {
        'depart_list': depart_list,
        # 分完页后的数据
        "page_queryset": page_queryset,
        # 生成的页码
        "pageNums": pageNums
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    """ 添加部门 """
    if request.method == 'POST':
        depart_name = request.POST.get('title')
        # 测试输出
        # print(depart_name)
        Department.objects.create(title=depart_name)

        # 部门添加成功后，重定向到部门列表页面
        return redirect('/depart/list/')

    return render(request, 'depart_add.html')


def depart_del(request):
    """ 删除部门 """
    # 获取部门ID
    nid = request.GET.get('nid')
    # 删除操作
    Department.objects.filter(id=nid).delete()

    return redirect('/depart/list')


def depart_edit(request, nid):
    """ 编辑部门 """
    if request.method == 'GET':
        # 根据nid获取部门信息，并获取Queryset对象中的第一条数据
        depart = Department.objects.filter(id=nid).first()
        # 测试输出
        # print(depart.title)
        # 将获取到的部门信息传到前端
        context = {
            'depart': depart
        }

        return render(request, 'depart_edit.html', context)

    elif request.method == 'POST':
        depart_name = request.POST.get('title')
        # 测试输出
        # print(depart_name)

        Department.objects.filter(id=nid).update(title=depart_name)

        return redirect('/depart/list/')


def depart_upload(request):
    """ 上传文件(Excel文件) —— 批量添加数据 """
    file_obj = request.FILES.get('exc')     # 获取用户上传的文件
    # file_name = file_obj.name           # 获取上传文件的名字

    # 对象传递给openpyxl，由openpyxl读取文件内容
    wb = load_workbook(file_obj)
    sheet = wb.worksheets[0]

    # 循环获取每一行的数据
    for i in sheet.iter_rows(min_row=2):
        text = i[0].value
        # print(text)     # 测试输出
        # 对提取出的部门进行去重
        exists = Department.objects.filter(title=text).exists()
        if not exists:
            Department.objects.create(title=text)

    return redirect('/depart/list/')









